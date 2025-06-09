import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pdfplumber
from typing import List, Tuple, Optional
from app.models.expense_model import ExpenseModel, ExpenseData

class ExpenseService:
    def __init__(self):
        self.model = ExpenseModel()

    def _extract_number_from_line(self, line: str) -> Optional[float]:
        """Trích xuất số từ một dòng văn bản"""
        try:
            parts = line.split()
            for part in reversed(parts):
                part_clean = part.replace(',', '')
                if part_clean.replace('.', '').replace('-', '').isdigit():
                    return float(part_clean)
        except (ValueError, IndexError):
            pass
        return None

    def _process_pdf_content(self, text: str) -> ExpenseData:
        """Xử lý nội dung PDF để trích xuất dữ liệu"""
        expense_data = ExpenseData()
        
        for line in text.split('\n'):
            line = line.strip()
            if "FREIGHT" in line and expense_data.air_freight == 0:
                if value := self._extract_number_from_line(line):
                    expense_data.air_freight = value
                    print(f"[PDF] AIR FREIGHT line: '{line}' | Số cuối: {value}")
                    
            elif "FUEL SURCHARGE" in line:
                if value := self._extract_number_from_line(line):
                    expense_data.fuel_surcharge = value
                    print(f"[PDF] FUEL SURCHARGE line: '{line}' | Số cuối: {value}")
                    
            elif "GRAND TOTAL" in line.upper():
                if value := self._extract_number_from_line(line):
                    expense_data.grand_total = value
                    print(f"[PDF] GRAND TOTAL line: '{line}' | Số cuối: {value}")
        
        expense_data.total = expense_data.air_freight + expense_data.fuel_surcharge
        return expense_data

    def extract_data_from_pdf(self, pdf_path: str) -> ExpenseData:
        """Trích xuất dữ liệu từ file PDF"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                expense_data = ExpenseData()
                for page in pdf.pages:
                    if text := page.extract_text():
                        page_data = self._process_pdf_content(text)
                        # Cập nhật dữ liệu từ trang hiện tại
                        expense_data.air_freight += page_data.air_freight
                        expense_data.fuel_surcharge += page_data.fuel_surcharge
                        expense_data.grand_total = max(expense_data.grand_total, page_data.grand_total)
                        expense_data.total = expense_data.air_freight + expense_data.fuel_surcharge
                
                print(f"[PDF] ==> TOTAL extracted from PDF: {expense_data.total}")
                return expense_data
        except Exception as e:
            print(f"Lỗi khi đọc file PDF {pdf_path}: {e}")
            return ExpenseData()

    def _validate_excel_columns(self, df: pd.DataFrame) -> None:
        """Kiểm tra các cột bắt buộc trong Excel"""
        if self.model.hbl_col not in df.columns:
            raise ValueError(f"Cột '{self.model.hbl_col}' không tồn tại trong file Excel")
        if self.model.expense_col not in df.columns:
            raise ValueError(f"Cột '{self.model.expense_col}' không tồn tại trong file Excel")

    def _update_excel_cell(self, worksheet: any, row: int, value: float, col_letter: str) -> None:
        """Cập nhật một ô trong Excel"""
        cell = f"{col_letter}{row}"
        worksheet[cell] = value
        print(f"[EXCEL] ==> Ghi giá trị {value} vào ô {cell}")

    def update_excel_with_pdf_data(self, excel_path: str, pdf_paths: List[str]) -> str:
        """Cập nhật file Excel với dữ liệu từ các file PDF"""
        try:
            # Đọc và validate Excel
            df = pd.read_excel(excel_path, header=4)
            self._validate_excel_columns(df)
            
            # Chuẩn bị workbook
            wb = load_workbook(excel_path)
            ws = wb.active
            expense_col_letter = get_column_letter(df.columns.get_loc(self.model.expense_col) + 1)
            header_offset = 4

            # Xử lý từng file PDF
            for pdf_path in pdf_paths:
                pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
                if pdf_name.startswith(tuple(f"{i}." for i in range(1, 100))):
                    pdf_name = pdf_name.split('.', 1)[1].strip()
                print(f"Processing PDF: {pdf_name}")

                # Xử lý từng dòng trong Excel
                for index, row in df.iterrows():
                    hbl_value = str(row[self.model.hbl_col]).strip()
                    print(f"Excel row {index}: HBL='{hbl_value}', PDF name='{pdf_name}'")

                    if pdf_name.upper().startswith("DEBIT NOTE"):
                        if hbl_value in pdf_name:
                            self._update_excel_cell(ws, index + header_offset + 2, 0, expense_col_letter)
                        continue

                    if pdf_name.startswith(hbl_value):
                        expense_data = self.extract_data_from_pdf(pdf_path)
                        self.model.set_expense_data(expense_data)

                        next_idx = index + 1
                        if next_idx < len(df):
                            next_row = df.iloc[next_idx]
                            if str(next_row[self.model.hbl_col]).strip() == "Freight":
                                freight_row = index + header_offset + 3
                                self._update_excel_cell(ws, freight_row, expense_data.total, expense_col_letter)

                                if next_idx + 1 < len(df):
                                    korea_row = df.iloc[next_idx + 1]
                                    if str(korea_row[self.model.hbl_col]).strip() == "Korea local cost":
                                        korea_row_num = freight_row + 1
                                        self._update_excel_cell(ws, korea_row_num, expense_data.korea_local_cost, expense_col_letter)
                        break

            # Lưu file
            output_dir = os.path.join(os.getcwd(), 'TONG_HOP_EXPENSE')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, "UPDATED_" + os.path.basename(excel_path))
            wb.save(output_path)
            return output_path

        except Exception as e:
            print(f"Lỗi khi cập nhật Excel: {e}")
            if 'df' in locals():
                print(f"Các cột hiện tại: {df.columns.tolist()}")
            raise
