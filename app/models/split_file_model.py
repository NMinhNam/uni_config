import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from copy import copy
from datetime import datetime

class SplitModel:
    def __init__(self):
        self.width_multiplier = 2.0
        self.row_height = 80
        self.data_start_row = 20
        self.footer_start_row = 34
        self.merge_columns = [
            "No.", "Bill No.", "CD No.", "INV", "AMOUNT",
            "SHIPPER - TERM", "Ex Rate", "CD Date. \n(Ngày tờ khai)",
            "CONT/LCL/AIR", "QUANTITY"
        ]

    def read_excel_file(self, file_path, sheet_name=None, header=0):
        """Read Excel file and return DataFrame"""
        return pd.read_excel(file_path, sheet_name=sheet_name, header=header)

    def load_workbooks(self, input_file_path, template_file_path, sheet_name=None):
        """Load workbooks and return their worksheet objects"""
        try:
            wb_input = load_workbook(input_file_path)
            ws_input = wb_input[sheet_name] if sheet_name else wb_input.active
            wb_template = load_workbook(template_file_path)
            ws_template = wb_template.active
            return wb_input, ws_input, wb_template, ws_template
        except Exception as e:
            print(f"Error loading workbooks: {str(e)}")
            return None, None, None, None

    def copy_cell_style(self, source_cell, target_cell):
        """Copy cell style from source to target"""
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font) if source_cell.font else None
            target_cell.fill = copy(source_cell.fill) if source_cell.fill else None
            target_cell.border = copy(source_cell.border) if source_cell.border else None
            target_cell.alignment = copy(source_cell.alignment) if source_cell.alignment else None
            target_cell.number_format = copy(source_cell.number_format) if source_cell.number_format else None

    def copy_dimensions(self, source_ws, target_ws, max_row, max_col):
        for col in range(1, max_col + 1):
            target_ws.column_dimensions[get_column_letter(col)].width = source_ws.column_dimensions[get_column_letter(col)].width
        for row in range(1, max_row + 1):
            target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

    def copy_header(self, ws_template, ws):
        for row in range(1, 20):
            for col in range(1, ws_template.max_column + 1):
                source_cell = ws_template.cell(row=row, column=col)
                target_cell = ws.cell(row=row, column=col)
                target_cell.value = source_cell.value
                self.copy_cell_style(source_cell, target_cell)

        for merged_range in ws_template.merged_cells.ranges:
            if merged_range.min_row <= 19:
                ws.merge_cells(start_row=merged_range.min_row, start_column=merged_range.min_col,
                              end_row=merged_range.max_row, end_column=merged_range.max_col)

        self.copy_dimensions(ws_template, ws, 19, ws_template.max_column)

    def copy_data(self, group, ws_input, ws, data_start_row):
        """Copy dữ liệu từ group vào worksheet mới"""
        group_indices = group.index.tolist()
        for row_idx, (group_idx, row) in enumerate(group.iterrows(), start=data_start_row):
            # Tính toán input_row dựa trên index thực tế của group
            input_row = group_idx + 4  # header=2 + offset 2
            ws.row_dimensions[row_idx].height = self.row_height
            
            # Debug logs
            print(f"Debug - Processing row {row_idx}, input_row: {input_row}")
            print(f"Debug - Group index: {group_idx}")
            
            for col_idx, value in enumerate(row, start=1):
                target_cell = ws.cell(row=row_idx, column=col_idx)
                target_cell.value = value
                
                # Chỉ copy style khi input_row hợp lệ
                if input_row <= ws_input.max_row and col_idx <= ws_input.max_column and input_row > 4:
                    try:
                        source_cell = ws_input.cell(row=input_row, column=col_idx)
                        self.copy_cell_style(source_cell, target_cell)
                        target_cell.font = Font(size=26)
                        
                        # Set column width cho dòng đầu tiên
                        if row_idx == data_start_row:
                            source_col_letter = get_column_letter(col_idx)
                            target_col_letter = get_column_letter(col_idx)
                            if source_col_letter in ws_input.column_dimensions:
                                original_width = ws_input.column_dimensions[source_col_letter].width
                                ws.column_dimensions[target_col_letter].width = original_width * self.width_multiplier
                    except Exception as e:
                        print(f"Warning - Could not copy cell style at row {input_row}, col {col_idx}: {str(e)}")
                        continue
                
        return group_indices

    def copy_merged_cells(self, ws_input, ws, group_indices, data_start_row):
        for merged_range in ws_input.merged_cells.ranges:
            if merged_range.min_row >= 4 and merged_range.min_row <= ws_input.max_row:
                source_row_start = merged_range.min_row
                source_row_end = merged_range.max_row
                if any(source_row_start <= group_idx + 4 <= source_row_end for group_idx in group_indices):
                    target_row_start = data_start_row + [group_idx + 4 for group_idx in group_indices].index(source_row_start)
                    target_row_end = target_row_start + (source_row_end - source_row_start)
                    ws.merge_cells(
                        start_row=target_row_start, 
                        start_column=merged_range.min_col,
                        end_row=target_row_end, 
                        end_column=merged_range.max_col
                    )

    def copy_footer(self, ws_template, ws, footer_start_row, output_start_row):
        if footer_start_row <= ws_template.max_row:
            for row in range(footer_start_row, ws_template.max_row + 1):
                target_row = output_start_row + (row - footer_start_row)
                for col in range(1, ws_template.max_column + 1):
                    source_cell = ws_template.cell(row=row, column=col)
                    target_cell = ws.cell(row=target_row, column=col)
                    is_merged = False
                    for merged_range in ws_template.merged_cells.ranges:
                        if (merged_range.min_row >= footer_start_row and
                            row >= merged_range.min_row and row <= merged_range.max_row and
                            col >= merged_range.min_col and col <= merged_range.max_col):
                            is_merged = True
                            if row == merged_range.min_row and col == merged_range.min_col:
                                top_left_cell = ws_template.cell(row=merged_range.min_row, column=merged_range.min_col)
                                target_cell.value = top_left_cell.value if top_left_cell.value is not None else ""
                                self.copy_cell_style(top_left_cell, target_cell)
                                ws.merge_cells(start_row=target_row, start_column=merged_range.min_col,
                                              end_row=target_row + (merged_range.max_row - merged_range.min_row),
                                              end_column=merged_range.max_col)
                            break
                    if not is_merged and not isinstance(source_cell, MergedCell):
                        target_cell.value = source_cell.value if source_cell.value is not None else ""
                        self.copy_cell_style(source_cell, target_cell)

    def split_file(self, input_file_path, template_file_path, output_dir):
        """Split an Excel file into multiple files based on a column (Client) using a template"""
        try:
            # Read the Excel file (default to first sheet)
            df = pd.read_excel(input_file_path, header=0)

            # Check if 'Client' column exists
            if 'Client' not in df.columns:
                return {'status': 'error', 'message': "Không tìm thấy cột 'Client' trong file Excel"}

            # Load workbooks
            wb_input = load_workbook(input_file_path)
            ws_input = wb_input.active
            wb_template = load_workbook(template_file_path)
            ws_template = wb_template.active

            # Group by 'Client'
            grouped = df.groupby('Client')
            processed_files = []

            # Generate a timestamp for file naming
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            # Process each group
            for client, group in grouped:
                # Create a new file for each client
                client_clean = str(client).strip().replace(' ', '_').replace('/', '_')
                file_name = f"{client_clean}_{timestamp}.xlsx"
                file_path = os.path.join(output_dir, file_name)

                # Create new workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Split"

                # Copy header from template
                self.copy_header(ws_template, ws)

                # Copy data for the group
                group_indices = self.copy_data(group, ws_input, ws, self.data_start_row)
                self.copy_merged_cells(ws_input, ws, group_indices, self.data_start_row)

                # Copy footer from template
                output_start_row = ws.max_row + 1
                self.copy_footer(ws_template, ws, self.footer_start_row, output_start_row)

                # Center align all data cells
                for row in ws.iter_rows(min_row=self.data_start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = Alignment(vertical='center', horizontal='center')

                # Save the file
                wb.save(file_path)
                processed_files.append(file_name)

            if not processed_files:
                return {'status': 'error', 'message': 'Không có dữ liệu để tách'}

            return {
                'status': 'success',
                'message': 'Đã tách file thành công',
                'files': processed_files
            }

        except Exception as e:
            return {'status': 'error', 'message': f'Lỗi khi xử lý file: {str(e)}'}

    def process_sheet(self, sheet_name, input_file_path, template_file_path, output_dir):
        """Process a single sheet from the Excel file"""
        try:
            # Đọc file Excel với header=2 như code gốc
            df = pd.read_excel(input_file_path, sheet_name=sheet_name, header=2)

            # Load workbooks theo đúng sheet
            wb_input = load_workbook(input_file_path)
            ws_input = wb_input[sheet_name]  # Lấy đúng sheet cần xử lý
            wb_template = load_workbook(template_file_path)
            ws_template = wb_template.active

            # Tìm cột Bill No.
            so_van_don_col = None
            for col in df.columns:
                col_clean = str(col).lower().replace(" ", "")
                if "bill" in col_clean and "no" in col_clean:
                    so_van_don_col = col
                    break

            if not so_van_don_col:
                return {
                    'status': 'error',
                    'message': f"Không tìm thấy cột 'Bill No.' trong sheet {sheet_name}"
                }

            # Forward fill giá trị Bill No.
            df[so_van_don_col] = df[so_van_don_col].ffill()
            df_all = df.copy()

            processed_files = []
            for van_don, group in df_all.groupby(so_van_don_col):
                # Tạo tên file theo định dạng gốc
                file_name = f"{str(van_don).strip().replace('/', '_').replace(' ', '')}_{sheet_name}.xlsx"
                file_path = os.path.join(output_dir, file_name)

                wb = Workbook()
                ws = wb.active
                ws.title = "CONT"

                # Copy cấu trúc từ template
                self.copy_header(ws_template, ws)
                group_indices = self.copy_data(group, ws_input, ws, self.data_start_row)
                self.copy_merged_cells(ws_input, ws, group_indices, self.data_start_row)
                
                # Merge cells theo cột được chỉ định
                total_row = self.merge_cells(ws, group, self.merge_columns, self.data_start_row)

                # Căn giữa tất cả các ô dữ liệu
                for row in ws.iter_rows(min_row=self.data_start_row, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(vertical='center', horizontal='center')

                # Copy footer
                output_start_row = ws.max_row + 1
                self.copy_footer(ws_template, ws, self.footer_start_row, output_start_row)

                # Lưu file
                wb.save(file_path)
                processed_files.append(file_name)

            return {
                'status': 'success',
                'message': f"Đã xử lý sheet {sheet_name}",
                'files': processed_files
            }

        except ValueError as e:
            if "Worksheet named" in str(e):
                print(f"Sheet '{sheet_name}' không tồn tại trong file")
                return {
                    'status': 'error',
                    'message': f"Sheet '{sheet_name}' không tồn tại trong file"
                }
            raise
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {str(e)}")
            return {
                'status': 'error',
                'message': f"Lỗi xử lý sheet {sheet_name}: {str(e)}"
            }
    
    def merge_cells(self, ws, group, merge_columns, data_start_row):
        """Merge cells based on specific columns"""
        # Tìm dòng TOTAL
        total_row = None
        for row_idx in range(data_start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                value = str(ws.cell(row=row_idx, column=col).value).upper().strip()
                if value == "TOTAL":
                    total_row = row_idx
                    break
            if total_row:
                break

        if total_row is None:
            total_row = ws.max_row + 1

        # Tìm cột Ex Rate
        ex_rate_col = None
        for col_idx, header in enumerate(group.columns, 1):
            header_clean = str(header).lower().replace(" ", "").replace("\n", "")
            if "exrate" in header_clean:
                ex_rate_col = col_idx
                break

        # Xử lý ngày trong Ex Rate
        if ex_rate_col and total_row:
            date_value = ws.cell(row=total_row, column=ex_rate_col).value
            if date_value:
                try:
                    if isinstance(date_value, str):
                        date_obj = datetime.strptime(date_value, "%m/%d/%Y")
                    elif isinstance(date_value, datetime):
                        date_obj = date_value
                    else:
                        raise ValueError("Giá trị ngày không hợp lệ")
                    formatted_date = date_obj.strftime("%m/%d/%Y")
                    target_cell = ws.cell(row=16, column=8)
                    target_cell.value = formatted_date
                    target_cell.alignment = Alignment(vertical='center', horizontal='center')
                    target_cell.font = Font(size=36)
                except Exception as e:
                    print(f"Lỗi khi định dạng ngày: {e}")

        # Merge các cột được chỉ định
        for merge_col in merge_columns:
            for col_idx, actual_col in enumerate(group.columns, 1):
                if merge_col.lower().replace(" ", "") == actual_col.lower().replace(" ", ""):
                    start_row = data_start_row
                    end_row = total_row - 1
                    if end_row >= start_row:
                        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                        cell = ws.cell(row=start_row, column=col_idx)
                        cell.alignment = Alignment(vertical='center', horizontal='center')
                    break

        # Merge dòng TOTAL
        if total_row:
            start_col = end_col = None
            for col_idx, header in enumerate(group.columns, 1):
                if header.lower().replace(" ", "") == "no.":
                    start_col = col_idx
                if header.lower().replace(" ", "") == "quantity":
                    end_col = col_idx
            if start_col and end_col and start_col <= end_col:
                ws.merge_cells(start_row=total_row, start_column=start_col, end_row=total_row, end_column=end_col)
                cell = ws.cell(row=total_row, column=start_col)
                cell.value = "TOTAL"
                cell.alignment = Alignment(vertical='center', horizontal='center')
                cell.font = Font(size=26)

        return total_row