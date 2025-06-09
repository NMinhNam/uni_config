import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pdfplumber
from dataclasses import dataclass
from typing import Optional, Dict, Any

@dataclass
class ExpenseData:
    """Class chứa dữ liệu từ PDF"""
    air_freight: float = 0.0
    fuel_surcharge: float = 0.0
    grand_total: float = 0.0
    total: float = 0.0

    @property
    def korea_local_cost(self) -> float:
        """Tính Korea local cost từ grand total và total"""
        return self.grand_total - self.total

class PdfExcelModel:
    def __init__(self):
        self.expense_col = "Expense" 
        self.hbl_col = "HBL"  

    def extract_total_from_pdf(self, pdf_path):
        """Trích xuất TOTAL từ file PDF dựa trên AIR FREIGHT, FUEL SURCHARGE hoặc GRAND TOTAL."""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                total = 0.0
                grand_total = 0.0
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    lines = text.split('\n')
                    air_freight = 0.0
                    fuel_surcharge = 0.0

                    for line in lines:
                        print(f"[PDF DEBUG] line: '{line}'")
                        line = line.strip()
                        if "FREIGHT" in line:
                            try:
                                parts = line.split()
                                for part in reversed(parts):
                                    part_clean = part.replace(',', '')
                                    if part_clean.replace('.', '').replace('-', '').isdigit():
                                        air_freight = float(part_clean)
                                        print(f"  [PDF] AIR FREIGHT line: '{line}' | Số cuối: {air_freight}")
                                        break
                            except (ValueError, IndexError):
                                continue
                        elif "FUEL SURCHARGE" in line:
                            try:
                                parts = line.split()
                                for part in reversed(parts):
                                    part_clean = part.replace(',', '')
                                    if part_clean.replace('.', '').replace('-', '').isdigit():
                                        fuel_surcharge = float(part_clean)
                                        print(f"  [PDF] FUEL SURCHARGE line: '{line}' | Số cuối: {fuel_surcharge}")
                                        break
                            except (ValueError, IndexError):
                                continue
                        # Luôn tìm GRAND TOTAL
                        if "GRAND TOTAL" in line.upper():
                            parts = line.replace(',', '').split()
                            for part in reversed(parts):
                                try:
                                    part_clean = part.replace(',', '')
                                    grand_total = float(part_clean)
                                    print(f"  [PDF] GRAND TOTAL line: '{line}' | Số cuối: {grand_total}")
                                    break
                                except ValueError:
                                    continue

                    total = air_freight + fuel_surcharge
                    print(f"  [PDF] ==> TOTAL extracted from PDF: {total}")

                return (total if total > 0 else None, grand_total)
        except Exception as e:
            print(f"Lỗi khi đọc file PDF {pdf_path}: {e}")
            return (None, 0.0)

    def update_excel_with_pdf_data(self, excel_path, pdf_paths):
        """Cập nhật file Excel với dữ liệu từ danh sách file PDF."""
        try:
            # Đọc file Excel, lấy index header
            df = pd.read_excel(excel_path, header=4)
            print(f"Các cột trong file Excel: {df.columns.tolist()}")

            if self.hbl_col not in df.columns:
                raise ValueError(f"Cột '{self.hbl_col}' không tồn tại trong file Excel. Vui lòng kiểm tra file Excel.")

            
            if self.expense_col not in df.columns:
                raise ValueError(f"Cột '{self.expense_col}' không tồn tại trong file Excel. Vui lòng kiểm tra file Excel.")

            wb = load_workbook(excel_path)
            ws = wb.active

            
            expense_col_idx = df.columns.get_loc(self.expense_col) + 1  # pandas index bắt đầu từ 0
            expense_col_letter = get_column_letter(expense_col_idx)

            header_offset = 4  # Số dòng header (header=4)
            for pdf_path in pdf_paths:
                pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
                
                if pdf_name.startswith(tuple(f"{i}." for i in range(1, 100))):
                    pdf_name = pdf_name.split('.', 1)[1].strip()
                print(f"Processing PDF: {pdf_name}")
                for index, row in df.iterrows():
                    hbl_value = str(row[self.hbl_col]).strip()
                    print(f"Excel row {index}: HBL='{hbl_value}', PDF name='{pdf_name}'")
                    if pdf_name.upper().startswith("DEBIT NOTE"):
                        for index, row in df.iterrows():
                            hbl_value = str(row[self.hbl_col]).strip()
                            if hbl_value in pdf_name:
                                hbl_excel_row = index + header_offset + 2
                                ws[f"{expense_col_letter}{hbl_excel_row}"] = 0
                                print(f"  [EXCEL] ==> Ghi 0 vào Excel: tại dòng {hbl_excel_row} (DEBIT NOTE)")
                                break
                        continue  
                    if pdf_name.startswith(hbl_value):
                        
                        if pdf_name.upper().startswith("DEBIT NOTE"):
                            hbl_excel_row = index + header_offset + 2
                            ws[f"{expense_col_letter}{hbl_excel_row}"] = 0
                            print(f"  [EXCEL] ==> Ghi 0 vào Excel: tại dòng {hbl_excel_row} (DEBIT NOTE)")
                            break
                        total, grand_total = self.extract_total_from_pdf(pdf_path)
                        print(f"  -> Match found! Extracted total from PDF: {total}, GRAND TOTAL: {grand_total}")
                        if total is not None:
                            
                            next_idx = index + 1
                            if next_idx < len(df):
                                next_row = df.iloc[next_idx]
                                next_hbl = str(next_row[self.hbl_col]).strip()
                                if next_hbl == "Freight":
                                    
                                    hbl_excel_row = index + header_offset + 2
                                    freight_excel_row = hbl_excel_row + 1
                                    print(f"  [DEBUG] Mã HBL ở dòng {hbl_excel_row} trong Excel")
                                    print(f"  [DEBUG] Dòng Freight ở dòng {freight_excel_row} trong Excel")
                                    ws[f"{expense_col_letter}{freight_excel_row}"] = total
                                    print(f"  [EXCEL] ==> Ghi TOTAL vào Excel: {total} tại dòng {freight_excel_row} (Freight)")

                                    
                                    if next_idx + 1 < len(df):
                                        korea_row = df.iloc[next_idx + 1]
                                        korea_hbl = str(korea_row[self.hbl_col]).strip()
                                        if korea_hbl == "Korea local cost":
                                            korea_excel_row = freight_excel_row + 1
                                            korea_local_cost = grand_total - total
                                            print(f"  [DEBUG] GRAND TOTAL lấy được: {grand_total}")
                                            print(f"  [DEBUG] TOTAL Freight lấy được: {total}")
                                            print(f"  [DEBUG] Korea local cost tính ra: {korea_local_cost}")
                                            ws[f"{expense_col_letter}{korea_excel_row}"] = korea_local_cost
                                            print(f"  [EXCEL] ==> Ghi Korea local cost vào Excel: {korea_local_cost} tại dòng {korea_excel_row}")
                            break  

            
            output_dir = os.path.join(os.getcwd(), 'TONG_HOP_EXPENSE')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, "UPDATED_" + os.path.basename(excel_path))
            wb.save(output_path)
            return output_path
        except Exception as e:
            print(f"Lỗi khi cập nhật Excel: {e}")
            print(f"Các cột hiện tại: {df.columns.tolist()}")
            raise

@dataclass
class ExpenseModel:
    """Model chứa thông tin và cấu trúc dữ liệu expense"""
    expense_col: str = "Expense"
    hbl_col: str = "HBL"
    expense_data: Optional[ExpenseData] = None
    
    def set_expense_data(self, data: ExpenseData) -> None:
        """Cập nhật dữ liệu expense"""
        self.expense_data = data
    
    def to_dict(self) -> Dict[str, Any]:
        """Chuyển đổi model thành dictionary"""
        return {
            'expense_col': self.expense_col,
            'hbl_col': self.hbl_col,
            'expense_data': {
                'air_freight': self.expense_data.air_freight if self.expense_data else 0.0,
                'fuel_surcharge': self.expense_data.fuel_surcharge if self.expense_data else 0.0,
                'grand_total': self.expense_data.grand_total if self.expense_data else 0.0,
                'total': self.expense_data.total if self.expense_data else 0.0
            } if self.expense_data else None
        }
    
    @staticmethod
    def from_dict(data: Dict[str, Any]) -> 'ExpenseModel':
        """Tạo model từ dictionary"""
        model = ExpenseModel(
            expense_col=data.get('expense_col', "Expense"),
            hbl_col=data.get('hbl_col', "HBL")
        )
        if expense_data := data.get('expense_data'):
            model.expense_data = ExpenseData(
                air_freight=expense_data.get('air_freight', 0.0),
                fuel_surcharge=expense_data.get('fuel_surcharge', 0.0),
                grand_total=expense_data.get('grand_total', 0.0),
                total=expense_data.get('total', 0.0)
            )
        return model